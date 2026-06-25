from netaiops_asset.chat_v2.followup import try_handle_v2_followup_analysis_forced
from netaiops_asset.chat_v2.semantic_router import build_v2_semantic_route, semantic_confirm_question_from_route
import io
import os
import tempfile
import time
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi import Request
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel

from netaiops_asset.agent.conversation_actions import detect_conversation_action, handle_conversation_action
from netaiops_asset.agent.conversation_store import (
    append_turn,
    create_conversation,
    delete_conversation,
    get_conversation,
    list_conversations,
)
from netaiops_asset.agent.rule_parser import parse_question
from netaiops_asset.cmdb.adapter import CMDBAdapter
from netaiops_asset.cmdb.field_map import CMDB_FIELDS, field_labels, normalize_fields
from netaiops_asset.config_loader import CONFIG_PATH, get_config
from netaiops_asset.llm.client import LLMClient
from netaiops_asset.llm.tool_planner import apply_llm_plan, plan_with_llm
from netaiops_asset.llm.planner_policy import accept_llm_parse, build_planner_diagnostics, should_try_llm
from netaiops_asset.security.audit import write_audit
from netaiops_asset.security.request_context import reset_request_context, set_request_context
from netaiops_asset.tools.cmdb_tools import CMDB_TOOL_CATALOG, tool_query_cmdb_device_detail, tool_query_cmdb_devices, tool_query_cmdb_devices_by_ips
from netaiops_asset.web.ui import render_index_html
from netaiops_asset.chat_v2.router import try_handle_v2_chat
from netaiops_asset.chat_v2.confirmation import store_pending_commands, try_handle_v2_execution_confirmation
from netaiops_asset.chat_v2.context import save_v2_context_from_response, get_context_debug
from netaiops_asset.chat_v2.followup import try_handle_v2_followup_analysis
from netaiops_asset.chat_v2.llm_intent_planner import planner_debug_payload
from netaiops_asset.chat_v2.plan_dispatcher import build_dispatch_debug_payload
from netaiops_asset.chat_v2.execution_response_enricher import normalize_execution_confirmation_question, enrich_v2_execution_response


CONFIG = get_config()
APP_NAME = CONFIG.get("app", {}).get("name", "netaiops-asset-agent")
APP_VERSION = CONFIG.get("app", {}).get("version", "0.1.0-v1-batch17-conversation-history")
START_TIME = time.time()

app = FastAPI(title=APP_NAME, version=APP_VERSION)

# V3_SHADOW_MODE_MARKER_BEGIN
def _v3_shadow_enabled():
    import os as _v3_shadow_os
    value = _v3_shadow_os.environ.get("NETAIOPS_V3_SHADOW_ENABLED", "1").strip().lower()
    return value not in {"0", "false", "no", "off", "disabled"}


def _v3_shadow_build(question, user=None, conversation_id=None, payload=None):
    if not _v3_shadow_enabled():
        return {"enabled": False, "decision": None, "plan": None, "error": "shadow_disabled"}

    question_text = str(question or "").strip()
    if not question_text:
        return {"enabled": False, "decision": None, "plan": None, "error": "empty_question"}

    try:
        from netaiops_asset.chat_v3.intent_arbiter import decide_intent
        from netaiops_asset.chat_v3.intent_dispatcher import build_dispatch_plan

        context = {
            "shadow_mode": True,
            "conversation_id": conversation_id or "",
            "payload_keys": sorted(list((payload or {}).keys())) if isinstance(payload, dict) else [],
        }

        decision = decide_intent(
            question=question_text,
            context=context,
            user=user,
            conversation_id=conversation_id,
        )

        plan = build_dispatch_plan(
            question=question_text,
            decision=decision,
            context=context,
            user=user,
            conversation_id=conversation_id,
        )

        return {
            "enabled": True,
            "decision": decision,
            "plan": plan,
            "error": "",
        }
    except Exception as exc:
        return {
            "enabled": True,
            "decision": None,
            "plan": None,
            "error": repr(exc),
        }


def _v3_shadow_response_summary(response):
    if not isinstance(response, dict):
        return {"response_type": type(response).__name__}

    v2_payload = response.get("v2") if isinstance(response.get("v2"), dict) else {}
    parsed_payload = response.get("parsed") if isinstance(response.get("parsed"), dict) else {}

    return {
        "response_type": "dict",
        "keys": sorted([str(key) for key in response.keys()])[:80],
        "status": response.get("status"),
        "planner_source": response.get("planner_source"),
        "request_id": response.get("request_id"),
        "conversation_id": response.get("conversation_id"),
        "count": response.get("count"),
        "returned": response.get("returned"),
        "parsed_intent": parsed_payload.get("intent"),
        "v2_keys": sorted([str(key) for key in v2_payload.keys()])[:80],
        "v2_execution_policy": v2_payload.get("execution_policy"),
    }


def _v3_shadow_write(shadow_state, question, user, conversation_id, v2_route, v2_response=None, extra=None):
    try:
        if not isinstance(shadow_state, dict):
            return
        if not shadow_state.get("enabled"):
            return

        import os as _v3_shadow_os
        from netaiops_asset.chat_v3.shadow_logger import write_shadow_record

        merged_extra = {
            "shadow_error": shadow_state.get("error", ""),
        }
        if isinstance(extra, dict):
            merged_extra.update(extra)

        # V3_PLAN_DECISION_NORMALIZATION_MARKER_BEGIN
        def _v3_local_as_dict(_value):
            if _value is None:
                return {}
            if isinstance(_value, dict):
                return dict(_value)
            try:
                from dataclasses import asdict as _v3_dc_asdict, is_dataclass as _v3_dc_is_dataclass
                if _v3_dc_is_dataclass(_value):
                    _data = _v3_dc_asdict(_value)
                    if isinstance(_data, dict):
                        return dict(_data)
            except Exception:
                pass
            for _method_name in ("as_dict", "model_dump", "dict"):
                _method = getattr(_value, _method_name, None)
                if callable(_method):
                    try:
                        _data = _method()
                        if isinstance(_data, dict):
                            return dict(_data)
                    except Exception:
                        pass
            try:
                _data = vars(_value)
                if isinstance(_data, dict):
                    return dict(_data)
            except Exception:
                pass
            _known_fields = (
                "action",
                "handler_key",
                "response_mode",
                "reason",
                "confidence",
                "effective_confidence",
                "accepted",
                "requires_confirmation",
                "safety_allowed",
                "answer",
                "items",
                "columns",
                "field_labels",
                "commands",
                "risks",
                "target",
                "device",
                "device_name",
                "intent",
            )
            _data = {}
            for _field in _known_fields:
                try:
                    if hasattr(_value, _field):
                        _data[_field] = getattr(_value, _field)
                except Exception:
                    pass
            if _data:
                return _data
            try:
                _data = dict(_value)
                if isinstance(_data, dict):
                    return _data
            except Exception:
                pass
            return {}

        try:
            _v3_normalized_plan = _v3_local_as_dict(shadow_state.get("plan"))
            _v3_normalized_decision = _v3_local_as_dict(shadow_state.get("decision"))
            if not isinstance(_v3_normalized_plan, dict):
                _v3_normalized_plan = {}
            if not isinstance(_v3_normalized_decision, dict):
                _v3_normalized_decision = {}
        except Exception as _v3_plan_normalization_exc:
            _v3_normalized_plan = {}
            _v3_normalized_decision = {}
            merged_extra["plan_decision_normalization_error"] = repr(_v3_plan_normalization_exc)
        # V3_PLAN_DECISION_NORMALIZATION_MARKER_END

        # V3_TAKEOVER_DRY_RUN_MARKER_BEGIN
        try:
            from netaiops_asset.chat_v3.takeover_gate import evaluate_takeover

            takeover_gate_runtime = evaluate_takeover(
                plan=_v3_normalized_plan,
                decision=_v3_normalized_decision,
            ).as_dict()
            takeover_gate_if_enabled = evaluate_takeover(
                plan=_v3_normalized_plan,
                decision=_v3_normalized_decision,
                enabled=True,
            ).as_dict()

            merged_extra["takeover_gate_runtime"] = takeover_gate_runtime
            merged_extra["takeover_gate_if_enabled"] = takeover_gate_if_enabled
        except Exception as _v3_takeover_gate_exc:
            merged_extra["takeover_gate_error"] = repr(_v3_takeover_gate_exc)
        # V3_TAKEOVER_DRY_RUN_MARKER_END

        # V3_RESPONSE_READINESS_DRY_RUN_MARKER_BEGIN
        try:
            from netaiops_asset.chat_v3.takeover_response import evaluate_response_readiness

            _v3_takeover_gate_for_response = merged_extra.get("takeover_gate_if_enabled") or merged_extra.get("takeover_gate_runtime") or {}
            response_readiness_if_enabled = evaluate_response_readiness(
                plan=_v3_normalized_plan,
                decision=_v3_normalized_decision,
                gate=_v3_takeover_gate_for_response,
            ).as_dict()
            merged_extra["takeover_response_readiness_if_enabled"] = response_readiness_if_enabled
        except Exception as _v3_response_readiness_exc:
            merged_extra["takeover_response_readiness_error"] = repr(_v3_response_readiness_exc)
        # V3_RESPONSE_READINESS_DRY_RUN_MARKER_END

        # V3_RESPONSE_GENERATOR_DRY_RUN_MARKER_BEGIN
        try:
            import os as _v3_response_generator_os
            from netaiops_asset.chat_v3.response_generator import generate_v3_response
            from netaiops_asset.chat_v3.takeover_response import evaluate_response_readiness

            _v3_live_llm_enabled = str(
                _v3_response_generator_os.getenv("NETAIOPS_V3_RESPONSE_GENERATOR_LIVE_LLM", "0")
            ).strip().lower() in {"1", "true", "yes", "on", "enabled"}
            _v3_gate_for_generator = merged_extra.get("takeover_gate_if_enabled") or merged_extra.get("takeover_gate_runtime") or {}
            _v3_generator_context = {
                "question": str(question or ""),
                "v2_route": v2_route,
                "v2_response": v2_response,
                "extra": merged_extra,
            }
            _v3_generated_response = generate_v3_response(
                question=str(question or ""),
                conversation_id=conversation_id,
                plan=_v3_normalized_plan,
                decision=_v3_normalized_decision,
                context=_v3_generator_context,
                gate=_v3_gate_for_generator,
                allow_live_llm=_v3_live_llm_enabled,
            ).as_dict()
            merged_extra["response_generator_runtime"] = _v3_generated_response

            if _v3_generated_response.get("ready") and _v3_generated_response.get("answer"):
                _v3_plan_after_generator = dict(_v3_normalized_plan or {})
                _v3_plan_after_generator["answer"] = _v3_generated_response.get("answer")
                if isinstance(_v3_generated_response.get("items"), list):
                    _v3_plan_after_generator["items"] = _v3_generated_response.get("items")
                if isinstance(_v3_generated_response.get("columns"), list):
                    _v3_plan_after_generator["columns"] = _v3_generated_response.get("columns")
                if isinstance(_v3_generated_response.get("field_labels"), dict):
                    _v3_plan_after_generator["field_labels"] = _v3_generated_response.get("field_labels")
                merged_extra["takeover_response_readiness_after_generator"] = evaluate_response_readiness(
                    plan=_v3_plan_after_generator,
                    decision=_v3_normalized_decision,
                    gate=_v3_gate_for_generator,
                ).as_dict()
        except Exception as _v3_response_generator_exc:
            merged_extra["response_generator_error"] = repr(_v3_response_generator_exc)
        # V3_RESPONSE_GENERATOR_DRY_RUN_MARKER_END

        write_shadow_record(
            question=str(question or ""),
            conversation_id=conversation_id,
            user=user,
            v2_route=str(v2_route or ""),
            v2_summary=_v3_shadow_response_summary(v2_response),
            v3_decision=_v3_normalized_decision,
            v3_plan=_v3_normalized_plan,
            shadow_dir=_v3_shadow_os.environ.get(
                "NETAIOPS_V3_SHADOW_DIR",
                "/var/lib/netaiops-asset-agent/data/v3_intent_shadow",
            ),
            extra=merged_extra,
        )
    except Exception:
        pass
# V3_SHADOW_MODE_MARKER_END



@app.middleware("http")
async def v2_chat_router_middleware(request, call_next):
    v3_shadow_state = None
    payload = {}
    question = ""
    user = None
    conversation_id = None
    if request.method == "POST" and request.url.path == "/api/v1/chat":
        try:
            payload = await request.json()
            question = str(payload.get("question") or "").strip()
            user = payload.get("user")
            conversation_id = payload.get("conversation_id")
            v3_shadow_state = _v3_shadow_build(
                question=question,
                user=user,
                conversation_id=conversation_id,
                payload=payload,
            )
            if question:
                # batch67_advice_analysis_route
                batch67_advice_response = _batch67_try_handle_advice_analysis(locals())
                if batch67_advice_response is not None:
                    _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "v2_advice_analysis", batch67_advice_response)
                    return JSONResponse(batch67_advice_response)
                # batch58_semantic_route_main
                # batch63_inline_command_execution
                # 用户输入中直接包含“立即执行 show/display 命令”时，优先提取并执行这些只读命令，
                # 避免被后续 follow-up 分析分支误判为“继续分析上一批结果”。
                try:
                    inline_response = try_handle_v2_inline_command_execution(
                        question=question,
                        conversation_id=conversation_id,
                        user=user,
                    )
                    if inline_response:
                        inline_response = _batch66_sync_history_before_return(locals(), inline_response)
                        _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "v2_inline_command_execute", inline_response)
                        return JSONResponse(inline_response)
                except Exception as _batch63_inline_exc:
                    try:
                        _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "v2_inline_command_execute_error", None, extra={"inline_error": repr(_batch63_inline_exc)})
                        return JSONResponse({
                            "status": "error",
                            "planner_source": "v2_inline_command_execution",
                            "answer": "内联命令执行分支异常：{}".format(repr(_batch63_inline_exc)),
                            "items": [],
                            "v2": {"inline_error": repr(_batch63_inline_exc)},
                        })
                    except Exception:
                        pass

                semantic_decision = build_v2_semantic_route(
                    question,
                    user=user,
                    conversation_id=conversation_id,
                )
                semantic_route = semantic_decision.get('route')
                if semantic_route == 'v2_execution_confirmation':
                    semantic_confirm_question = semantic_confirm_question_from_route(question, semantic_decision)
                    confirm_response = try_handle_v2_execution_confirmation(
                        semantic_confirm_question,
                        user=user,
                        conversation_id=conversation_id,
                    )
                    confirm_response = enrich_v2_execution_response(
                        confirm_response,
                        question=question,
                        user=user,
                        conversation_id=conversation_id,
                    )
                    if isinstance(confirm_response, dict):
                        confirm_response.setdefault('v2', {})
                        confirm_response['v2']['semantic_decision'] = semantic_decision
                    if confirm_response:
                        # batch60_persist_execution_context: execution result must be persisted before returning.
                        try:
                            _batch60_cid = conversation_id or confirm_response.get('conversation_id')
                            if _batch60_cid:
                                confirm_response['conversation_id'] = _batch60_cid
                            save_v2_context_from_response(
                                conversation_id=_batch60_cid,
                                user=user,
                                question=question,
                                response=confirm_response,
                            )
                        except Exception as _batch60_exc:
                            try:
                                confirm_response.setdefault('v2', {})
                                confirm_response['v2']['context_save_error'] = repr(_batch60_exc)
                            except Exception:
                                pass
                        confirm_response = _batch66_sync_history_before_return(locals(), confirm_response)
                        _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "v2_semantic_execution_confirmation", confirm_response)
                        return JSONResponse(confirm_response)
                elif semantic_route == 'v2_followup_analysis':
                    followup_response = try_handle_v2_followup_analysis_forced(
                        question,
                        user=user,
                        conversation_id=conversation_id,
                    )
                    if isinstance(followup_response, dict):
                        followup_response.setdefault('v2', {})
                        followup_response['v2']['semantic_decision'] = semantic_decision
                    if followup_response is not None:
                        followup_response = _batch66_sync_history_before_return(locals(), followup_response)
                        _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "v2_semantic_followup_analysis", followup_response)
                        return JSONResponse(followup_response)

                # batch57_execution_priority_pre_followup
                if is_v2_execution_request_question(question):
                    normalized_confirm_question = normalize_execution_confirmation_question(question)
                    confirm_response = try_handle_v2_execution_confirmation(
                        normalized_confirm_question,
                        user=user,
                        conversation_id=conversation_id,
                    )
                    confirm_response = enrich_v2_execution_response(
                        confirm_response,
                        question=question,
                        user=user,
                        conversation_id=conversation_id,
                    )
                    if confirm_response:
                        # batch60_persist_execution_context: execution result must be persisted before returning.
                        try:
                            _batch60_cid = conversation_id or confirm_response.get('conversation_id')
                            if _batch60_cid:
                                confirm_response['conversation_id'] = _batch60_cid
                            save_v2_context_from_response(
                                conversation_id=_batch60_cid,
                                user=user,
                                question=question,
                                response=confirm_response,
                            )
                        except Exception as _batch60_exc:
                            try:
                                confirm_response.setdefault('v2', {})
                                confirm_response['v2']['context_save_error'] = repr(_batch60_exc)
                            except Exception:
                                pass
                        confirm_response = _batch66_sync_history_before_return(locals(), confirm_response)
                        _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "v2_execution_request_confirmation", confirm_response)
                        return JSONResponse(confirm_response)

                followup_response = try_handle_v2_followup_analysis(
                    question,
                    user=user,
                    conversation_id=conversation_id,
                )
                if followup_response is not None:
                    request_id = write_audit({
                        "user": user,
                        "question": question,
                        "intent": (followup_response.get("parsed") or {}).get("intent", "v2_followup_analysis"),
                        "tool_name": "v2_followup_analysis",
                        "tool_args": {
                            "parsed": followup_response.get("parsed"),
                        },
                        "data_source": "v2_conversation_context",
                        "result_count": followup_response.get("count", 0),
                        "returned_count": followup_response.get("returned", 0),
                        "status": followup_response.get("status", "unknown"),
                    })
                    followup_response["request_id"] = request_id
                    cid, _ = append_turn(conversation_id, question, followup_response, user=user)
                    followup_response["conversation_id"] = cid
                    try:
                        save_v2_context_from_response(
                            conversation_id=cid,
                            user=user,
                            question=question,
                            response=followup_response,
                        )
                    except Exception:
                        pass
                    followup_response = _batch66_sync_history_before_return(locals(), followup_response)
                    _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "v2_followup_analysis", followup_response)
                    return JSONResponse(followup_response)

                normalized_confirm_question = normalize_execution_confirmation_question(question)
                confirm_response = try_handle_v2_execution_confirmation(
                    normalized_confirm_question,
                    user=user,
                    conversation_id=conversation_id,
                )
                confirm_response = enrich_v2_execution_response(
                    confirm_response,
                    question=question,
                    user=user,
                    conversation_id=conversation_id,
                )
                if confirm_response is not None:
                    request_id = write_audit({
                        "user": user,
                        "question": question,
                        "intent": (confirm_response.get("parsed") or {}).get("intent", "v2_execute_confirmation"),
                        "tool_name": "v2_execution_confirmation",
                        "tool_args": {
                            "parsed": confirm_response.get("parsed"),
                        },
                        "data_source": "netmiko_mcp_confirmed_execution",
                        "result_count": confirm_response.get("count", 0),
                        "returned_count": confirm_response.get("returned", 0),
                        "status": confirm_response.get("status", "unknown"),
                    })
                    confirm_response["request_id"] = request_id
                    cid, _ = append_turn(conversation_id, question, confirm_response, user=user)
                    confirm_response["conversation_id"] = cid
                    try:
                        save_v2_context_from_response(
                            conversation_id=cid,
                            user=user,
                            question=question,
                            response=confirm_response,
                        )
                    except Exception:
                        pass
                    confirm_response = _batch66_sync_history_before_return(locals(), confirm_response)
                    _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "v2_execution_confirmation", confirm_response)
                    return JSONResponse(confirm_response)

                v2_response = try_handle_v2_chat(question, user=user, conversation_id=conversation_id)
                if v2_response is not None:
                    request_id = write_audit({
                        "user": user,
                        "question": question,
                        "intent": (v2_response.get("parsed") or {}).get("intent", "v2_troubleshoot"),
                        "tool_name": "v2_chat_router_middleware",
                        "tool_args": {
                            "parsed": v2_response.get("parsed"),
                            "execution_policy": (v2_response.get("v2") or {}).get("execution_policy"),
                        },
                        "data_source": "cmdb_netmiko_prometheus_v2_router",
                        "result_count": v2_response.get("count", 0),
                        "returned_count": v2_response.get("returned", 0),
                        "status": v2_response.get("status", "unknown"),
                    })
                    v2_response["request_id"] = request_id
                    cid, _ = append_turn(conversation_id, question, v2_response, user=user)
                    v2_response["conversation_id"] = cid
                    try:
                        store_pending_commands(
                            conversation_id=cid,
                            user=user,
                            question=question,
                            response=v2_response,
                        )
                    except Exception:
                        pass
                    try:
                        save_v2_context_from_response(
                            conversation_id=cid,
                            user=user,
                            question=question,
                            response=v2_response,
                        )
                    except Exception:
                        pass
                    v2_response = _batch66_sync_history_before_return(locals(), v2_response)
                    _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "v2_chat_router", v2_response)
                    return JSONResponse(v2_response)
        except Exception as exc:
            # Do not break existing V1 chat flow if V2 router fails.
            try:
                write_audit({
                    "user": None,
                    "question": "",
                    "intent": "v2_router_middleware_error",
                    "tool_name": "v2_chat_router_middleware",
                    "tool_args": {},
                    "data_source": "v2_router",
                    "result_count": 0,
                    "returned_count": 0,
                    "status": "failed",
                    "error": repr(exc),
                })
            except Exception:
                pass

    _v3_shadow_write(v3_shadow_state, question, user, conversation_id, "call_next_to_legacy_chat", None)
    return await call_next(request)


@app.middleware("http")
async def request_context_middleware(request: Request, call_next):
    forwarded_for = request.headers.get("x-forwarded-for", "")
    if forwarded_for:
        client_ip = forwarded_for.split(",")[0].strip()
    else:
        client_ip = request.client.host if request.client else ""

    token = set_request_context({
        "client_ip": client_ip,
        "user_agent": request.headers.get("user-agent", ""),
        "method": request.method,
        "path": request.url.path,
    })

    try:
        response = await call_next(request)
        return response
    finally:
        reset_request_context(token)


class ChatRequest(BaseModel):
    question: str
    user: str | None = "local_user"
    limit: int | None = 20
    conversation_id: str | None = None
    planner_mode: str | None = "llm"
    debug: bool | None = False


class ConversationCreateRequest(BaseModel):
    title: str | None = None
    user: str | None = "web_user"


class LLMParseRequest(BaseModel):
    question: str
    user: str | None = "web_user"
    force: bool | None = True

class LLMCompareRequest(BaseModel):
    question: str
    user: str | None = "web_user"
    planner_mode: str | None = "llm"

class ToolQueryDevicesRequest(BaseModel):
    filters: dict[str, Any] | None = None
    fields: list[str] | str | None = None
    page: int | None = 1
    page_size: int | None = 20
    user: str | None = "tool_user"


class ToolQueryDeviceDetailRequest(BaseModel):
    keyword: str
    fields: list[str] | str | None = None
    user: str | None = "tool_user"


class ToolQueryDevicesByIpsRequest(BaseModel):
    ips: list[str]
    fields: list[str] | str | None = None
    page_size: int | None = None
    user: str | None = "tool_user"


def build_answer(result: dict[str, Any]) -> str:
    if result.get("status") != "ok":
        return f"CMDB 查询失败：{result.get('message', '未知错误')}"

    count = int(result.get("count") or 0)
    returned = int(result.get("returned") or 0)
    filters = result.get("filters", {})

    if count == 0:
        return "未查询到符合条件的 CMDB 网络设备记录。请确认 IDC、机房、机柜、主机名、管理IP 或型号是否正确。"

    limit_note = ""
    if count > returned:
        limit_note = f" 当前仅展示前 {returned} 条，可缩小查询条件；如需离线查看，可导出 Excel，单次最多导出 500 条。"

    if filters:
        filter_text = "，".join([f"{k}={v}" for k, v in filters.items()])
        return f"根据基金 CMDB 网络设备查询条件 {filter_text}，共查询到 {count} 条记录，本次返回 {returned} 条。{limit_note}"

    return f"基金 CMDB 网络设备查询完成，共查询到 {count} 条记录，本次返回 {returned} 条。{limit_note}"


@app.get("/", response_class=HTMLResponse)
def index_page() -> str:
    return render_index_html()


@app.get("/ui", response_class=HTMLResponse)
def ui_page() -> str:
    return render_index_html()


@app.get("/health")
def health() -> dict[str, Any]:
    config = get_config()
    cmdb = config.get("cmdb", {})
    return {
        "status": "ok",
        "service": APP_NAME,
        "version": APP_VERSION,
        "config_path": CONFIG_PATH,
        "uptime_seconds": int(time.time() - START_TIME),
        "v1_scope": "fund_cmdb_network_server_query",
        "cmdb_mode": cmdb.get("mode", "not_configured"),
        "cmdb_env": cmdb.get("env"),
        "cmdb_base_url": cmdb.get("api_base_url"),
        "features": config.get("features", {}),
        "llm_enabled": bool(config.get("llm", {}).get("enabled", False)),
        "llm_model": config.get("llm", {}).get("model"),
    }


@app.get("/api/v1/cmdb/schema")
def cmdb_schema() -> dict[str, Any]:
    return {
        "status": "ok",
        "data_source": "fund_cmdb_networkServer",
        "mode": get_config().get("cmdb", {}).get("mode", "not_configured"),
        "fields": CMDB_FIELDS,
        "field_labels": field_labels(),
    }


@app.get("/api/v1/cmdb/probe")
def cmdb_probe() -> dict[str, Any]:
    adapter = CMDBAdapter()
    result = adapter.probe()
    request_id = write_audit({
        "user": "api_user",
        "question": "api_cmdb_probe",
        "intent": "probe",
        "tool_name": "cmdb_probe",
        "tool_args": {},
        "data_source": "fund_cmdb_networkServer",
        "result_count": result.get("count", 0),
        "returned_count": result.get("returned", 0),
        "status": result.get("status", "unknown"),
    })
    result["request_id"] = request_id
    return result


@app.get("/api/v1/cmdb/devices")
def query_devices(
    search: str | None = Query(None),
    IDC: str | None = Query(None),
    server_room: str | None = Query(None),
    rack: str | None = Query(None),
    host_name: str | None = Query(None),
    mgmt_ip: str | None = Query(None),
    sn: str | None = Query(None),
    ci_type: str | None = Query(None),
    manufacturer: str | None = Query(None),
    band: str | None = Query(None),
    device_spec: str | None = Query(None),
    os_version: str | None = Query(None),
    env: str | None = Query(None),
    status: str | None = Query(None),
    tag: str | None = Query(None),
    maintenance_manufacturer: str | None = Query(None),
    fields: str | None = Query(None),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=100),
) -> dict[str, Any]:
    filters = {
        "search": search,
        "IDC__icontains": IDC,
        "server_room__icontains": server_room,
        "rack__icontains": rack,
        "host_name__icontains": host_name,
        "mgmt_ip": mgmt_ip,
        "sn__icontains": sn,
        "ci_type__icontains": ci_type,
        "manufacturer__icontains": manufacturer,
        "band__icontains": band,
        "device_spec__icontains": device_spec,
        "os_version__icontains": os_version,
        "env": env,
        "status__icontains": status,
        "tag__icontains": tag,
        "maintenance_manufacturer__icontains": maintenance_manufacturer,
    }
    filters = {k: v for k, v in filters.items() if v not in (None, "")}

    adapter = CMDBAdapter()
    result = adapter.query_devices(filters=filters, fields=fields, page=page, page_size=pageSize)
    request_id = write_audit({
        "user": "api_user",
        "question": "api_query_devices",
        "intent": "query_devices",
        "tool_name": "query_devices",
        "tool_args": {"filters": filters, "fields": fields, "page": page, "pageSize": pageSize},
        "data_source": "fund_cmdb_networkServer",
        "result_count": result.get("count", 0),
        "returned_count": result.get("returned", 0),
        "status": result.get("status", "unknown"),
    })
    result["request_id"] = request_id
    result["answer"] = build_answer(result)
    return result


@app.get("/api/v1/cmdb/device/detail")
def query_device_detail(
    keyword: str = Query(..., description="host_name, mgmt_ip, server_ID or sn"),
    fields: str | None = Query(None),
) -> dict[str, Any]:
    adapter = CMDBAdapter()
    result = adapter.query_device_detail(keyword=keyword, fields=fields)
    request_id = write_audit({
        "user": "api_user",
        "question": "api_query_device_detail",
        "intent": "query_device_detail",
        "tool_name": "query_device_detail",
        "tool_args": {"keyword": keyword, "fields": fields},
        "data_source": "fund_cmdb_networkServer",
        "result_count": result.get("count", 0),
        "returned_count": result.get("returned", 0),
        "status": result.get("status", "unknown"),
    })
    result["request_id"] = request_id
    result["answer"] = build_answer(result)
    return result


@app.get("/api/v1/cmdb/devices/by-ips")
def query_devices_by_ips(
    ips: str = Query(..., description="Comma, whitespace or newline separated management IPs"),
    fields: str | None = Query(None),
    pageSize: int = Query(100, ge=1, le=100),
) -> dict[str, Any]:
    raw_items = []
    for part in ips.replace("\n", ",").replace("\r", ",").replace(" ", ",").replace("，", ",").split(","):
        item = part.strip()
        if item:
            raw_items.append(item)

    ip_list = []
    for item in raw_items:
        if item not in ip_list:
            ip_list.append(item)

    if not ip_list:
        result = {"status": "error", "message": "ips is empty", "count": 0, "returned": 0, "items": []}
    else:
        adapter = CMDBAdapter()
        ip_param = ",".join(ip_list)
        result = adapter.query_devices(
            filters={"mgmt_ip__in": ip_param},
            fields=fields,
            page=1,
            page_size=min(max(pageSize, len(ip_list)), 100),
        )

    request_id = write_audit({
        "user": "api_user",
        "question": "api_query_devices_by_ips",
        "intent": "query_devices_by_ips",
        "tool_name": "query_devices_by_ips",
        "tool_args": {"ips_count": len(ip_list), "fields": fields, "pageSize": pageSize},
        "data_source": "fund_cmdb_networkServer",
        "result_count": result.get("count", 0),
        "returned_count": result.get("returned", 0),
        "status": result.get("status", "unknown"),
    })
    result["request_id"] = request_id
    result["answer"] = build_answer(result)
    return result


@app.get("/api/v1/cmdb/explore/mgmt-ip-list")
def explore_mgmt_ip_list(ip: str = Query(...)) -> dict[str, Any]:
    adapter = CMDBAdapter()
    result = adapter.explore_mgmt_ip_list(ip)
    request_id = write_audit({
        "user": "api_user",
        "question": "api_explore_mgmt_ip_list",
        "intent": "explore_mgmt_ip_list",
        "tool_name": "explore_mgmt_ip_list",
        "tool_args": {"ip": ip},
        "data_source": "fund_cmdb_networkServer",
        "result_count": 0,
        "returned_count": 0,
        "status": result.get("status", "unknown"),
    })
    result["request_id"] = request_id
    return result



@app.get("/api/v1/cmdb/devices/export.xlsx")
def export_devices_xlsx(request: Request):
    import json
    import re as _re
    from urllib.parse import quote

    from fastapi import HTTPException
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from netaiops_asset.cmdb.field_map import QUERY_FILTER_FIELDS, normalize_field_name

    raw_params = dict(request.query_params)

    fields = raw_params.get("fields")

    try:
        page_size = int(raw_params.get("pageSize") or raw_params.get("page_size") or 500)
    except Exception:
        page_size = 500
    page_size = max(1, min(page_size, 500))

    query_alias = {
        "IDC": "IDC__icontains",
        "server_room": "server_room__icontains",
        "rack": "rack__icontains",
        "host_name": "host_name__icontains",
        "sn": "sn__icontains",
        "ci_type": "ci_type__icontains",
        "manufacturer": "manufacturer__icontains",
        "band": "band__icontains",
        "device_spec": "device_spec__icontains",
        "os_version": "os_version__icontains",
        "status": "status__icontains",
        "tag": "tag__icontains",
        "maintenance_manufacturer": "maintenance_manufacturer__icontains",
        "server_ID": "server_ID__icontains",
        "comment": "comment__icontains",
        "oa_contract": "oa_contract__icontains",
        "costcontrol_ticket_id": "costcontrol_ticket_id__icontains",
    }

    skip_keys = {"fields", "pageSize", "page_size", "page"}
    filters = {}

    for key, value in request.query_params.multi_items():
        if key in skip_keys or value in (None, ""):
            continue

        if key == "mgmt_ip":
            filters["mgmt_ip"] = value
            continue

        if key == "search":
            filters["search"] = value
            continue

        if key in query_alias:
            filters[query_alias[key]] = value
            continue

        normalized = normalize_field_name(key)
        if normalized in QUERY_FILTER_FIELDS:
            filters[normalized] = value
            continue

        if "__" in normalized:
            base = normalized.split("__", 1)[0]
            if base in QUERY_FILTER_FIELDS:
                filters[normalized] = value
                continue

    adapter = CMDBAdapter()
    result = adapter.query_devices(filters=filters, fields=fields, page=1, page_size=page_size)

    if result.get("status") != "ok":
        raise HTTPException(
            status_code=502,
            detail={
                "message": "CMDB export query failed",
                "error_code": result.get("error_code"),
                "cmdb_message": result.get("message"),
                "http_status": result.get("http_status"),
            },
        )

    columns = result.get("fields") or normalize_fields(fields)
    labels = result.get("field_labels") or field_labels()
    rows = result.get("items", [])

    def cell_value(value):
        if value is None:
            return ""
        if isinstance(value, (list, dict)):
            return json.dumps(value, ensure_ascii=False)
        return value

    wb = Workbook()
    ws = wb.active
    ws.title = "CMDB网络设备"

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append([labels.get(c, c) for c in columns])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([cell_value(row.get(c, "")) for c in columns])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        label = labels.get(col, col)
        max_len = len(str(label))
        for row in rows[:200]:
            value = cell_value(row.get(col, ""))
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    request_id = write_audit({
        "user": "api_user",
        "question": "api_export_devices_xlsx",
        "intent": "export_devices_xlsx",
        "tool_name": "export_devices_xlsx",
        "tool_args": {"filters": filters, "fields": fields, "pageSize": page_size},
        "data_source": "fund_cmdb_networkServer",
        "result_count": result.get("count", 0),
        "returned_count": result.get("returned", 0),
        "status": result.get("status", "unknown"),
    })

    headers = {"Content-Disposition": f'attachment; filename="netaiops_cmdb_devices_{request_id}.xlsx"'}
    return StreamingResponse(
        iter([bio.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/api/v1/audit/recent")
def audit_recent(limit: int = Query(20, ge=1, le=200)) -> dict[str, Any]:
    import json

    config = get_config()
    audit_dir = Path(config.get("runtime", {}).get("audit_dir", "/var/lib/netaiops-asset-agent/data/audit"))

    items: list[dict[str, Any]] = []
    if audit_dir.exists():
        for f in sorted(audit_dir.glob("audit_*.jsonl"), reverse=True)[:10]:
            try:
                lines = f.read_text(encoding="utf-8").splitlines()
            except Exception:
                continue

            for line in reversed(lines):
                if not line.strip():
                    continue
                try:
                    event = json.loads(line)
                    items.append(event)
                except Exception:
                    continue
                if len(items) >= limit:
                    break
            if len(items) >= limit:
                break

    return {"status": "ok", "count": len(items), "items": items[:limit]}


@app.get("/api/v1/selfcheck")
def selfcheck() -> dict[str, Any]:
    config = get_config()
    cmdb = config.get("cmdb", {})
    runtime = config.get("runtime", {})

    audit_dir = Path(runtime.get("audit_dir", "/var/lib/netaiops-asset-agent/data/audit"))
    export_dir = Path(runtime.get("export_dir", "/var/lib/netaiops-asset-agent/data/exports"))
    conversation_dir = Path(runtime.get("conversation_dir", "/var/lib/netaiops-asset-agent/data/conversations"))

    token_env = cmdb.get("api_token_env", "NETAIOPS_CMDB_API_TOKEN")
    token_configured = bool(os.getenv(token_env, ""))

    def writable(path: Path) -> bool:
        try:
            path.mkdir(parents=True, exist_ok=True)
            with tempfile.NamedTemporaryFile("w", dir=str(path), delete=True, encoding="utf-8") as f:
                f.write("ok")
            return True
        except Exception:
            return False

    excel_ok = False
    try:
        import openpyxl  # noqa: F401
        excel_ok = True
    except Exception:
        excel_ok = False

    adapter = CMDBAdapter()
    probe = adapter.probe()
    cmdb_ok = probe.get("status") == "ok" and int(probe.get("http_status") or 0) == 200

    checks = {
        "config_loaded": bool(config),
        "token_configured": token_configured,
        "cmdb_api_reachable": cmdb_ok,
        "audit_dir_writable": writable(audit_dir),
        "export_dir_writable": writable(export_dir),
        "conversation_dir_writable": writable(conversation_dir),
        "excel_export_available": excel_ok,
    }

    overall = "ok" if all(checks.values()) else "warn"

    return {
        "status": overall,
        "service": APP_NAME,
        "version": APP_VERSION,
        "cmdb": {
            "mode": cmdb.get("mode"),
            "env": cmdb.get("env"),
            "api_base_url": cmdb.get("api_base_url"),
            "network_server_path": cmdb.get("network_server_path"),
            "token_env": token_env,
            "token_configured": token_configured,
        },
        "limits": config.get("limits", {}),
        "runtime": {
            "audit_dir": str(audit_dir),
            "export_dir": str(export_dir),
            "conversation_dir": str(conversation_dir),
            "audit_retention_days": runtime.get("audit_retention_days", 90),
            "conversation_retention_days": runtime.get("conversation_retention_days", 180),
        },
        "checks": checks,
        "cmdb_probe": {
            "status": probe.get("status"),
            "http_status": probe.get("http_status"),
            "count": probe.get("count"),
            "code": probe.get("code"),
            "message": probe.get("message"),
        },
    }



@app.get("/api/v1/llm/config")
def api_llm_config() -> dict[str, Any]:
    client = LLMClient()
    return {
        "status": "ok",
        "llm": client.masked_config(),
    }


@app.get("/api/v1/llm/models")
def api_llm_models() -> dict[str, Any]:
    client = LLMClient()
    result = client.list_models()
    write_audit({
        "user": "api_user",
        "question": "api_llm_models",
        "intent": "llm_models",
        "tool_name": "llm_models",
        "tool_args": {"model": client.model, "base_url": client.base_url},
        "data_source": "llm_gateway",
        "result_count": 1 if result.get("status") == "ok" else 0,
        "returned_count": 1 if result.get("status") == "ok" else 0,
        "status": result.get("status", "unknown"),
    })
    return result


@app.get("/api/v1/llm/probe")
def api_llm_probe() -> dict[str, Any]:
    client = LLMClient()
    result = client.probe()
    write_audit({
        "user": "api_user",
        "question": "api_llm_probe",
        "intent": "llm_probe",
        "tool_name": "llm_probe",
        "tool_args": {"model": client.model, "base_url": client.base_url},
        "data_source": "llm_gateway",
        "result_count": 1 if result.get("status") == "ok" else 0,
        "returned_count": 1 if result.get("status") == "ok" else 0,
        "status": result.get("status", "unknown"),
    })
    return result


@app.post("/api/v1/llm/parse")
def api_llm_parse(req: LLMParseRequest) -> dict[str, Any]:
    rule_parsed = parse_question(req.question)
    plan_result = plan_with_llm(req.question, rule_parsed=rule_parsed)
    parsed = apply_llm_plan(plan_result)

    write_audit({
        "user": req.user,
        "question": req.question,
        "intent": "llm_parse",
        "tool_name": "llm_tool_planner",
        "tool_args": {"rule_parsed": rule_parsed},
        "data_source": "llm_gateway",
        "result_count": 1 if plan_result.get("status") == "ok" else 0,
        "returned_count": 1 if parsed else 0,
        "status": plan_result.get("status", "unknown"),
    })

    return {
        "status": "ok" if parsed else "error",
        "question": req.question,
        "rule_parsed": rule_parsed,
        "llm_plan_result": plan_result,
        "parsed": parsed,
    }



@app.post("/api/v1/llm/compare")
def api_llm_compare(req: LLMCompareRequest) -> dict[str, Any]:
    v2_response = try_handle_v2_chat(req.question, user=req.user, conversation_id=req.conversation_id)
    if v2_response is not None:
        request_id = write_audit({
            "user": req.user,
            "question": req.question,
            "intent": (v2_response.get("parsed") or {}).get("intent", "v2_troubleshoot"),
            "tool_name": "v2_chat_router",
            "tool_args": {
                "parsed": v2_response.get("parsed"),
                "execution_policy": (v2_response.get("v2") or {}).get("execution_policy"),
            },
            "data_source": "cmdb_netmiko_prometheus_v2_router",
            "result_count": v2_response.get("count", 0),
            "returned_count": v2_response.get("returned", 0),
            "status": v2_response.get("status", "unknown"),
        })
        v2_response["request_id"] = request_id
        cid, _ = append_turn(req.conversation_id, req.question, v2_response, user=req.user)
        v2_response["conversation_id"] = cid
        try:
            store_pending_commands(
                conversation_id=cid,
                user=locals().get("user") or getattr(locals().get("req", None), "user", None),
                question=locals().get("question") or getattr(locals().get("req", None), "question", ""),
                response=v2_response,
            )
        except Exception:
            pass
        try:
            save_v2_context_from_response(
                conversation_id=cid,
                user=locals().get("user") or getattr(locals().get("req", None), "user", None),
                question=locals().get("question") or getattr(locals().get("req", None), "question", ""),
                response=v2_response,
            )
        except Exception:
            pass
        return v2_response

    llm_cfg = get_config().get("llm", {})
    rule_parsed = parse_question(req.question)

    should_try, should_try_reason = should_try_llm(
        req.question,
        rule_parsed,
        req.planner_mode,
        llm_cfg,
    )

    llm_plan_result = None
    llm_parsed = None
    accepted = False
    accept_reason = "llm_not_tried"
    selected_parsed = rule_parsed
    planner_source = "rule_parser"

    if should_try:
        llm_plan_result = plan_with_llm(req.question, rule_parsed=rule_parsed)
        llm_parsed = apply_llm_plan(llm_plan_result)
        accepted, accept_reason = accept_llm_parse(
            rule_parsed,
            llm_parsed,
            llm_plan_result,
            req.planner_mode,
            llm_cfg,
        )
        if accepted and llm_parsed:
            selected_parsed = llm_parsed
            planner_source = "llm_tool_planner"

    diagnostics = build_planner_diagnostics(
        requested_mode=req.planner_mode,
        should_try=should_try,
        should_try_reason=should_try_reason,
        accepted=accepted,
        accept_reason=accept_reason,
        rule_parsed=rule_parsed,
        llm_plan_result=llm_plan_result,
    )

    write_audit({
        "user": req.user,
        "question": req.question,
        "intent": "llm_compare",
        "tool_name": "llm_tool_planner",
        "tool_args": {"planner_mode": req.planner_mode, "diagnostics": diagnostics},
        "data_source": "llm_gateway",
        "result_count": 1 if llm_plan_result and llm_plan_result.get("status") == "ok" else 0,
        "returned_count": 1 if selected_parsed else 0,
        "status": "ok",
    })

    return {
        "status": "ok",
        "question": req.question,
        "planner_source": planner_source,
        "diagnostics": diagnostics,
        "rule_parsed": rule_parsed,
        "llm_plan_result": llm_plan_result,
        "llm_parsed": llm_parsed,
        "selected_parsed": selected_parsed,
    }


@app.get("/api/v1/conversations")
def api_list_conversations(limit: int = Query(50, ge=1, le=200)) -> dict[str, Any]:
    items = list_conversations(limit=limit)
    return {"status": "ok", "count": len(items), "items": items}


@app.post("/api/v1/conversations")
def api_create_conversation(req: ConversationCreateRequest) -> dict[str, Any]:
    conv = create_conversation(title=req.title or "新对话", user=req.user)
    request_id = write_audit({
        "user": req.user or "web_user",
        "question": "api_create_conversation",
        "intent": "create_conversation",
        "tool_name": "create_conversation",
        "tool_args": {"conversation_id": conv.get("conversation_id")},
        "data_source": "conversation_store",
        "result_count": 1,
        "returned_count": 1,
        "status": "ok",
    })
    conv["request_id"] = request_id
    return {"status": "ok", "conversation": conv}


@app.get("/api/v1/conversations/{conversation_id}")
def api_get_conversation(conversation_id: str) -> dict[str, Any]:
    conv = get_conversation(conversation_id)
    if conv is None:
        raise HTTPException(status_code=404, detail="conversation not found")
    return {"status": "ok", "conversation": conv}


@app.delete("/api/v1/conversations/{conversation_id}")
def api_delete_conversation(conversation_id: str) -> dict[str, Any]:
    ok = delete_conversation(conversation_id)
    write_audit({
        "user": "web_user",
        "question": "api_delete_conversation",
        "intent": "delete_conversation",
        "tool_name": "delete_conversation",
        "tool_args": {"conversation_id": conversation_id},
        "data_source": "conversation_store",
        "result_count": 1 if ok else 0,
        "returned_count": 1 if ok else 0,
        "status": "ok" if ok else "not_found",
    })
    return {"status": "ok" if ok else "not_found", "deleted": ok}



@app.get("/api/v1/tools/catalog")
def api_tools_catalog() -> dict[str, Any]:
    return {
        "status": "ok",
        "count": len(CMDB_TOOL_CATALOG),
        "items": CMDB_TOOL_CATALOG,
    }


@app.post("/api/v1/tools/cmdb/query")
def api_tool_cmdb_query(req: ToolQueryDevicesRequest) -> dict[str, Any]:
    result = tool_query_cmdb_devices(
        filters=req.filters or {},
        fields=req.fields,
        page=req.page or 1,
        page_size=req.page_size or 20,
    )
    request_id = write_audit({
        "user": req.user,
        "question": "tool_query_cmdb_devices",
        "intent": "tool_query_cmdb_devices",
        "tool_name": "query_cmdb_devices",
        "tool_args": {"filters": req.filters or {}, "fields": req.fields, "page": req.page, "page_size": req.page_size},
        "data_source": "fund_cmdb_networkServer",
        "result_count": result.get("count", 0),
        "returned_count": result.get("returned", 0),
        "status": result.get("status", "unknown"),
    })
    result["request_id"] = request_id
    result["answer"] = build_answer(result)
    return result


@app.post("/api/v1/tools/cmdb/detail")
def api_tool_cmdb_detail(req: ToolQueryDeviceDetailRequest) -> dict[str, Any]:
    result = tool_query_cmdb_device_detail(
        keyword=req.keyword,
        fields=req.fields,
    )
    request_id = write_audit({
        "user": req.user,
        "question": "tool_query_cmdb_device_detail",
        "intent": "tool_query_cmdb_device_detail",
        "tool_name": "query_cmdb_device_detail",
        "tool_args": {"keyword": req.keyword, "fields": req.fields},
        "data_source": "fund_cmdb_networkServer",
        "result_count": result.get("count", 0),
        "returned_count": result.get("returned", 0),
        "status": result.get("status", "unknown"),
    })
    result["request_id"] = request_id
    result["answer"] = build_answer(result)
    return result


@app.post("/api/v1/tools/cmdb/query-by-ips")
def api_tool_cmdb_query_by_ips(req: ToolQueryDevicesByIpsRequest) -> dict[str, Any]:
    result = tool_query_cmdb_devices_by_ips(
        ips=req.ips,
        fields=req.fields,
        page_size=req.page_size,
    )
    request_id = write_audit({
        "user": req.user,
        "question": "tool_query_cmdb_devices_by_ips",
        "intent": "tool_query_cmdb_devices_by_ips",
        "tool_name": "query_cmdb_devices_by_ips",
        "tool_args": {"ips_count": len(req.ips or []), "fields": req.fields, "page_size": req.page_size},
        "data_source": "fund_cmdb_networkServer",
        "result_count": result.get("count", 0),
        "returned_count": result.get("returned", 0),
        "status": result.get("status", "unknown"),
    })
    result["request_id"] = request_id
    result["answer"] = build_answer(result)
    return result


@app.post("/api/v1/chat")
def chat(req: ChatRequest) -> dict[str, Any]:
    action_result = handle_conversation_action(req.question, req.conversation_id)
    if action_result is not None:
        request_id = write_audit({
            "user": req.user,
            "question": req.question,
            "intent": "conversation_action",
            "tool_name": action_result.get("action"),
            "tool_args": {
                "conversation_id": req.conversation_id,
                "export_params": action_result.get("export_params"),
                "source_turn_id": action_result.get("source_turn_id"),
            },
            "data_source": "conversation_store",
            "result_count": action_result.get("count", 0),
            "returned_count": action_result.get("returned", 0),
            "status": action_result.get("status", "unknown"),
        })

        response = {
            "status": action_result.get("status", "ok"),
            "request_id": request_id,
            "question": req.question,
            "parsed": {
                "intent": "conversation_action",
                "action": action_result.get("action"),
                "reason": "conversation_action_detector",
            },
            "llm_plan": None,
            "planner_source": "conversation_action",
            "planner_diagnostics": None,
            "action": action_result.get("action"),
            "answer": action_result.get("answer"),
            "columns": action_result.get("columns", []),
            "field_labels": action_result.get("field_labels", field_labels()),
            "count": action_result.get("count", 0),
            "returned": action_result.get("returned", 0),
            "items": action_result.get("items", []),
            "export_url": action_result.get("export_url"),
            "export_params": action_result.get("export_params"),
            "source_turn_id": action_result.get("source_turn_id"),
        }

        cid, _ = append_turn(req.conversation_id, req.question, response, user=req.user)
        response["conversation_id"] = cid
        return response

    llm_cfg = get_config().get("llm", {})

    rule_parsed = parse_question(req.question)
    parsed = rule_parsed

    llm_plan_result = None
    llm_parsed = None
    planner_source = "rule_parser"

    should_try, should_try_reason = should_try_llm(
        req.question,
        rule_parsed,
        req.planner_mode,
        llm_cfg,
    )

    accepted = False
    accept_reason = "llm_not_tried"

    if should_try:
        llm_plan_result = plan_with_llm(req.question, rule_parsed=rule_parsed)
        llm_parsed = apply_llm_plan(llm_plan_result)
        accepted, accept_reason = accept_llm_parse(
            rule_parsed,
            llm_parsed,
            llm_plan_result,
            req.planner_mode,
            llm_cfg,
        )

        if accepted and llm_parsed:
            parsed = llm_parsed
            planner_source = "llm_tool_planner"

    planner_diagnostics = build_planner_diagnostics(
        requested_mode=req.planner_mode,
        should_try=should_try,
        should_try_reason=should_try_reason,
        accepted=accepted,
        accept_reason=accept_reason,
        rule_parsed=rule_parsed,
        llm_plan_result=llm_plan_result,
    )

    if parsed.get("intent") == "clarify":
        request_id = write_audit({
            "user": req.user,
            "question": req.question,
            "intent": "clarify",
            "tool_name": "llm_tool_planner" if planner_source == "llm_tool_planner" else None,
            "tool_args": {"parsed": parsed, "planner_diagnostics": planner_diagnostics},
            "data_source": "llm_gateway" if planner_source == "llm_tool_planner" else None,
            "result_count": 0,
            "returned_count": 0,
            "status": "need_clarification",
        })
        response = {
            "status": "need_clarification",
            "request_id": request_id,
            "question": req.question,
            "parsed": parsed,
            "llm_plan": llm_plan_result,
            "planner_source": planner_source,
            "planner_diagnostics": planner_diagnostics if req.debug else None,
            "answer": parsed.get("message"),
            "items": [],
            "columns": [],
            "field_labels": field_labels(),
            "count": 0,
            "returned": 0,
        }
        cid, _ = append_turn(req.conversation_id, req.question, response, user=req.user)
        response["conversation_id"] = cid
        return response

    if parsed.get("intent") == "query_device_detail":
        result = tool_query_cmdb_device_detail(
            keyword=parsed.get("keyword", ""),
            fields=parsed.get("fields"),
        )
        tool_name = "query_cmdb_device_detail"
        tool_args = {"keyword": parsed.get("keyword"), "fields": parsed.get("fields")}
    else:
        limit = max(1, min(int(req.limit or 20), 100))
        result = tool_query_cmdb_devices(
            filters=parsed.get("filters", {}),
            fields=parsed.get("fields"),
            page=1,
            page_size=limit,
        )
        tool_name = "query_cmdb_devices"
        tool_args = {"filters": parsed.get("filters", {}), "fields": parsed.get("fields"), "limit": limit}

    request_id = write_audit({
        "user": req.user,
        "question": req.question,
        "intent": parsed.get("intent"),
        "tool_name": tool_name,
        "tool_args": {"tool_args": tool_args, "planner_source": planner_source, "planner_diagnostics": planner_diagnostics},
        "data_source": "fund_cmdb_networkServer",
        "result_count": result.get("count", 0),
        "returned_count": result.get("returned", 0),
        "status": result.get("status", "unknown"),
    })

    response = {
        "status": result.get("status", "ok"),
        "request_id": request_id,
        "question": req.question,
        "parsed": parsed,
        "llm_plan": llm_plan_result,
        "planner_source": planner_source,
        "planner_diagnostics": planner_diagnostics if req.debug else None,
        "answer": build_answer(result),
        "columns": result.get("fields", normalize_fields(None)),
        "field_labels": result.get("field_labels", field_labels()),
        "count": result.get("count", 0),
        "returned": result.get("returned", 0),
        "items": result.get("items", []),
    }

    cid, _ = append_turn(req.conversation_id, req.question, response, user=req.user)
    response["conversation_id"] = cid
    return response


@app.post("/api/v1/netmiko/validate_commands")
async def api_v1_netmiko_validate_commands(request: Request):
    """
    Validate Netmiko command suggestions.

    Safety:
    - This endpoint only validates command strings.
    - It does NOT execute any device CLI command.
    """
    from netaiops_asset.netmiko.cli_guard import CliReadOnlyGuard

    payload = await request.json()
    device_name = str(payload.get("device_name") or "").strip()
    device_type = str(payload.get("device_type") or "").strip()
    platform = str(payload.get("platform") or "").strip()
    commands = payload.get("commands")

    if commands is None:
        single = payload.get("command")
        commands = [single] if single else []

    if not isinstance(commands, list):
        commands = [commands]

    guard = CliReadOnlyGuard()
    items = []

    for idx, command in enumerate(commands, 1):
        command_text = str(command or "").strip()
        result = guard.validate(
            command_text,
            platform=platform or None,
            device_type=device_type or None,
        ).to_dict()

        items.append({
            "index": idx,
            "device_name": device_name,
            "device_type": device_type,
            "command": command_text,
            "guard": result,
            "can_execute_after_confirm": bool(result.get("status") == "passed"),
            "confirm_required": bool(result.get("status") == "passed"),
        })

    passed_count = sum(1 for x in items if (x.get("guard") or {}).get("status") == "passed")
    review_count = sum(1 for x in items if (x.get("guard") or {}).get("status") == "review")
    blocked_count = sum(1 for x in items if (x.get("guard") or {}).get("status") == "blocked")

    resp = {
        "status": "ok",
        "device_name": device_name,
        "device_type": device_type,
        "count": len(items),
        "passed_count": passed_count,
        "review_count": review_count,
        "blocked_count": blocked_count,
        "items": items,
    }

    try:
        request_id = write_audit({
            "user": payload.get("user"),
            "question": "netmiko_validate_commands",
            "intent": "v2_netmiko_validate_commands",
            "tool_name": "cli_guard",
            "tool_args": {
                "device_name": device_name,
                "device_type": device_type,
                "commands": commands,
            },
            "data_source": "local_cli_guard",
            "result_count": len(items),
            "returned_count": len(items),
            "status": "ok",
        })
        resp["request_id"] = request_id
    except Exception:
        pass

    resp = _batch66_sync_history_before_return(locals(), resp)
    return JSONResponse(resp)


@app.post("/api/v1/netmiko/execute_confirmed")
async def api_v1_netmiko_execute_confirmed(request: Request):
    """
    Execute one confirmed read-only Netmiko command.

    Required safety conditions:
    - CLI Guard status must be passed.
    - confirm_execute must equal YES.
    - Only send_command_and_get_output is used.
    """
    from netaiops_asset.netmiko.executor import ConfirmedNetmikoExecutor

    payload = await request.json()

    device_name = str(payload.get("device_name") or "").strip()
    command = str(payload.get("command") or "").strip()
    device_type = str(payload.get("device_type") or "").strip() or None
    platform = str(payload.get("platform") or "").strip() or None
    confirm_execute = str(payload.get("confirm_execute") or "").strip()
    confirmed_by = str(payload.get("confirmed_by") or payload.get("user") or "").strip() or None

    try:
        timeout = int(payload.get("timeout") or 60)
    except Exception:
        timeout = 60

    executor = ConfirmedNetmikoExecutor()
    result = executor.execute_confirmed(
        device_name=device_name,
        command=command,
        platform=platform,
        device_type=device_type,
        confirm_execute=confirm_execute,
        confirmed_by=confirmed_by,
        timeout=timeout,
    )

    try:
        request_id = write_audit({
            "user": confirmed_by,
            "question": "netmiko_execute_confirmed",
            "intent": "v2_netmiko_execute_confirmed",
            "tool_name": "confirmed_netmiko_executor",
            "tool_args": {
                "device_name": device_name,
                "device_type": device_type,
                "platform": platform,
                "command": command,
                "confirm_execute": confirm_execute,
            },
            "data_source": "netmiko_mcp",
            "result_count": 1 if result.get("ok") else 0,
            "returned_count": 1,
            "status": result.get("status"),
            "error": result.get("error"),
        })
        result["request_id"] = request_id
    except Exception:
        pass

    result = _batch66_sync_history_before_return(locals(), result)
    return JSONResponse(result)


@app.get("/api/v1/netmiko/safety_policy")
async def api_v1_netmiko_safety_policy():
    return JSONResponse({
        "status": "ok",
        "auto_execute": False,
        "execute_requires": [
            "CLI Guard status must be passed",
            "confirm_execute must equal YES",
            "confirmed_by should identify the operator",
            "Only send_command_and_get_output is allowed",
            "Config tool set_config_commands_and_commit_or_save is not exposed",
        ],
        "guard_status": {
            "passed": "clearly read-only, can enter confirmation flow",
            "review": "sensitive/high-output/active probing, not executable by this endpoint",
            "blocked": "dangerous or configuration command, rejected",
        },
    })


@app.get("/api/v1/v2/context")
async def api_v1_v2_context_debug(conversation_id: str = "", user: str = ""):
    """
    Debug endpoint for V2 conversation context.

    Safety:
    - Read-only local context JSON.
    - Does not execute device command.
    """
    return JSONResponse(get_context_debug(
        conversation_id=conversation_id or None,
        user=user or None,
    ))


@app.post("/api/v1/v2/llm_plan")
async def api_v1_v2_llm_plan(payload: dict):
    """
    Debug endpoint for V2 LLM-first intent planner.

    Safety:
    - Does not execute CLI.
    - Does not call Netmiko.
    - Only returns structured plan.
    """
    question = str(payload.get("question") or "")
    user = payload.get("user")
    conversation_id = payload.get("conversation_id")

    context = None
    try:
        context = get_context_debug(conversation_id=conversation_id or None, user=user or None).get("context")
    except Exception:
        context = None

    return JSONResponse(planner_debug_payload(question, context=context, user=user))


@app.post("/api/v1/v2/dispatch_plan")
async def api_v1_v2_dispatch_plan(payload: dict):
    """
    Debug endpoint for V2 Plan Validator + Action Dispatcher.

    Safety:
    - Does not execute CLI.
    - Does not call Netmiko.
    - Only returns plan + dispatch decision.
    """
    question = str(payload.get("question") or "")
    user = payload.get("user")
    conversation_id = payload.get("conversation_id")

    context = None
    try:
        context = get_context_debug(conversation_id=conversation_id or None, user=user or None).get("context")
    except Exception:
        context = None

    return JSONResponse(build_dispatch_debug_payload(question, context=context, user=user))
from netaiops_asset.chat_v2.execution_response_enricher import is_v2_execution_request_question
from netaiops_asset.chat_v2.inline_command_execute import try_handle_v2_inline_command_execution


@app.post("/api/v1/v2/semantic_route")
async def api_v1_v2_semantic_route(payload: dict):
    """
    Debug endpoint for unified V2 semantic route.

    Safety:
    - Does not execute CLI.
    - Does not call Netmiko.
    - Only returns LLM plan + dispatcher route.
    """
    question = str(payload.get("question") or "")
    user = payload.get("user")
    conversation_id = payload.get("conversation_id")

    decision = build_v2_semantic_route(
        question,
        user=user,
        conversation_id=conversation_id,
    )
    return JSONResponse({
        "status": "ok",
        "question": question,
        "decision": decision,
    })


# ===== Batch63-fix frontend markdown middleware =====
# 只对 text/html 响应动态注入前端 Markdown 渲染脚本。
# 不修改后端 answer 内容，不对 LLM 输出做二次概括。

try:
    from fastapi.responses import Response as _Batch63HTMLResponse
    from fastapi.staticfiles import StaticFiles as _Batch63StaticFiles

    try:
        if not any(getattr(r, "path", None) == "/static" for r in getattr(app, "routes", [])):
            app.mount("/static", _Batch63StaticFiles(directory="/opt/netaiops-asset-agent/static"), name="static")
    except Exception:
        pass

    @app.middleware("http")
    async def _batch63_markdown_html_inject_middleware(request, call_next):
        response = await call_next(request)

        try:
            content_type = response.headers.get("content-type", "")
            if "text/html" not in content_type.lower():
                return response

            body = b""
            async for chunk in response.body_iterator:
                body += chunk

            try:
                html = body.decode("utf-8")
            except Exception:
                return _Batch63HTMLResponse(
                    content=body,
                    status_code=response.status_code,
                    headers={k: v for k, v in response.headers.items() if k.lower() != "content-length"},
                    media_type=content_type,
                )

            script = '<script src="/static/batch63_markdown_render.js?v=batch64-safe-md"></script>'
            if "batch63_markdown_render.js" not in html:
                if "</body>" in html:
                    html = html.replace("</body>", script + "\n</body>", 1)
                elif "</html>" in html:
                    html = html.replace("</html>", script + "\n</html>", 1)
                else:
                    html = html + "\n" + script + "\n"

            headers = {k: v for k, v in response.headers.items() if k.lower() not in ("content-length", "content-encoding")}
            return _Batch63HTMLResponse(
                content=html,
                status_code=response.status_code,
                headers=headers,
                media_type="text/html",
            )
        except Exception:
            return response

except Exception:
    pass


# ===== Batch66 V2 history sync patch =====
def _batch66_extract_question_from_locals(_locals):
    try:
        for key in ("question", "user_question", "q", "query"):
            val = _locals.get(key)
            if isinstance(val, str) and val.strip():
                return val

        for obj in list(_locals.values()):
            if hasattr(obj, "question"):
                val = getattr(obj, "question", None)
                if isinstance(val, str) and val.strip():
                    return val
            if isinstance(obj, dict):
                val = obj.get("question")
                if isinstance(val, str) and val.strip():
                    return val
    except Exception:
        pass
    return ""


def _batch66_extract_user_from_locals(_locals):
    try:
        for key in ("user", "username", "user_id"):
            val = _locals.get(key)
            if isinstance(val, str) and val.strip():
                return val

        for obj in list(_locals.values()):
            if hasattr(obj, "user"):
                val = getattr(obj, "user", None)
                if isinstance(val, str) and val.strip():
                    return val
            if isinstance(obj, dict):
                val = obj.get("user")
                if isinstance(val, str) and val.strip():
                    return val
    except Exception:
        pass
    return "web_user"


def _batch66_extract_conversation_id_from_locals(_locals, response):
    try:
        if isinstance(response, dict) and response.get("conversation_id"):
            return response.get("conversation_id")

        for key in ("conversation_id", "cid"):
            val = _locals.get(key)
            if isinstance(val, str) and val.strip():
                return val

        for obj in list(_locals.values()):
            if hasattr(obj, "conversation_id"):
                val = getattr(obj, "conversation_id", None)
                if isinstance(val, str) and val.strip():
                    return val
            if isinstance(obj, dict):
                val = obj.get("conversation_id")
                if isinstance(val, str) and val.strip():
                    return val
    except Exception:
        pass
    return None


def _batch66_should_sync_history(response):
    if not isinstance(response, dict):
        return False

    planner = str(response.get("planner_source") or "")
    if planner in ("v2_execution_confirmation", "v2_followup_analysis", "v2_inline_command_execution"):
        return True

    v2 = response.get("v2")
    if isinstance(v2, dict):
        if v2.get("inline_command_execution"):
            return True
        if v2.get("batch65_bulk20_inline_execution"):
            return True
        if v2.get("batch63_fix5_yes_confirmation"):
            return True

    return False


def _batch66_call_append_turn(conversation_id, question, response, user):
    from netaiops_asset.agent.conversation_store import append_turn
    import inspect

    attempts = []

    call_specs = [
        ("kwargs_all", (), {"conversation_id": conversation_id, "question": question, "response": response, "user": user}),
        ("pos3_user_kw", (conversation_id, question, response), {"user": user}),
        ("pos4", (conversation_id, question, response, user), {}),
        ("pos3", (conversation_id, question, response), {}),
    ]

    for name, args, kwargs in call_specs:
        try:
            result = append_turn(*args, **kwargs)
            return {"ok": True, "method": name, "result": repr(result)[:500]}
        except TypeError as exc:
            attempts.append({"method": name, "error": repr(exc)[:500]})
        except Exception as exc:
            attempts.append({"method": name, "error": repr(exc)[:500]})
            break

    try:
        sig = inspect.signature(append_turn)
        kw = {}
        for pname in sig.parameters:
            low = pname.lower()
            if "conversation" in low or low in ("cid", "conversationid"):
                kw[pname] = conversation_id
            elif low in ("question", "query", "prompt", "user_question"):
                kw[pname] = question
            elif low in ("response", "answer", "result"):
                kw[pname] = response
            elif low in ("user", "username", "user_id"):
                kw[pname] = user
        if kw:
            result = append_turn(**kw)
            return {"ok": True, "method": "adaptive_kwargs", "result": repr(result)[:500]}
    except Exception as exc:
        attempts.append({"method": "adaptive_kwargs", "error": repr(exc)[:500]})

    return {"ok": False, "attempts": attempts[-8:]}


def _batch66_sync_history_before_return(_locals, response):
    try:
        if not _batch66_should_sync_history(response):
            return response

        if not isinstance(response, dict):
            return response

        v2 = response.setdefault("v2", {})
        if isinstance(v2, dict) and v2.get("batch66_history_synced"):
            return response

        question = _batch66_extract_question_from_locals(_locals)
        user = _batch66_extract_user_from_locals(_locals)
        conversation_id = _batch66_extract_conversation_id_from_locals(_locals, response)

        if not question or not conversation_id:
            if isinstance(v2, dict):
                v2["batch66_history_sync"] = {
                    "ok": False,
                    "reason": "missing_question_or_conversation_id",
                    "has_question": bool(question),
                    "has_conversation_id": bool(conversation_id),
                }
            return response

        response["conversation_id"] = conversation_id

        result = _batch66_call_append_turn(
            conversation_id=conversation_id,
            question=question,
            response=response,
            user=user,
        )

        if isinstance(v2, dict):
            v2["batch66_history_synced"] = bool(result.get("ok"))
            v2["batch66_history_sync"] = result

        return response

    except Exception as exc:
        try:
            if isinstance(response, dict):
                response.setdefault("v2", {})["batch66_history_sync_error"] = repr(exc)
        except Exception:
            pass
        return response


# ===== Batch67 advice analysis route patch =====
def _batch67_extract_question_from_locals(_locals):
    try:
        for key in ("question", "user_question", "q", "query"):
            val = _locals.get(key)
            if isinstance(val, str) and val.strip():
                return val
        for obj in list(_locals.values()):
            if hasattr(obj, "question"):
                val = getattr(obj, "question", None)
                if isinstance(val, str) and val.strip():
                    return val
            if isinstance(obj, dict):
                val = obj.get("question")
                if isinstance(val, str) and val.strip():
                    return val
    except Exception:
        pass
    return ""


def _batch67_extract_user_from_locals(_locals):
    try:
        for key in ("user", "username", "user_id"):
            val = _locals.get(key)
            if isinstance(val, str) and val.strip():
                return val
        for obj in list(_locals.values()):
            if hasattr(obj, "user"):
                val = getattr(obj, "user", None)
                if isinstance(val, str) and val.strip():
                    return val
            if isinstance(obj, dict):
                val = obj.get("user")
                if isinstance(val, str) and val.strip():
                    return val
    except Exception:
        pass
    return "web_user"


def _batch67_extract_conversation_id_from_locals(_locals):
    try:
        for key in ("conversation_id", "cid"):
            val = _locals.get(key)
            if isinstance(val, str) and val.strip():
                return val
        for obj in list(_locals.values()):
            if hasattr(obj, "conversation_id"):
                val = getattr(obj, "conversation_id", None)
                if isinstance(val, str) and val.strip():
                    return val
            if isinstance(obj, dict):
                val = obj.get("conversation_id")
                if isinstance(val, str) and val.strip():
                    return val
    except Exception:
        pass
    return None


def _batch67_contains_any(text, words):
    s = str(text or "").lower()
    return any(str(w).lower() in s for w in words)


def _batch67_is_advice_analysis_question(question):
    q = str(question or "").strip()
    if not q:
        return False

    hard_execute_phrases = [
        "确认可以执行",
        "确认执行",
        "确认在设备上执行",
        "确认执行全部命令 yes",
        "确认执行全部命令 YES",
        "执行以上命令",
        "直接在设备上执行",
        "立即执行命令",
        "帮我执行命令",
        "执行下面",
        "执行以下",
    ]
    if _batch67_contains_any(q, hard_execute_phrases):
        return False

    explicit_no_command = _batch67_contains_any(q, [
        "不要给命令",
        "不要生成命令",
        "不要执行设备",
        "不要执行命令",
        "只给建议",
        "只给操作建议",
        "只做分析",
        "单纯分析",
        "不用执行",
        "不需要执行",
    ])

    command_generation_phrases = [
        "给我命令",
        "给出命令",
        "查看命令",
        "排查命令",
        "只读命令",
        "show ",
        "display ",
        "怎么查",
        "查一下设备",
        "采集",
    ]

    if _batch67_contains_any(q, command_generation_phrases) and not explicit_no_command:
        if not _batch67_contains_any(q, ["是否推荐", "是否建议", "更推荐", "优缺点", "风险", "方案"]):
            return False

    advice_keywords = [
        "是否建议",
        "是否推荐",
        "更推荐",
        "建议先",
        "建议使用",
        "推荐使用",
        "优缺点",
        "利弊",
        "风险",
        "方案",
        "操作建议",
        "风险分析",
        "评估",
        "可行性",
        "是否合适",
        "是否合理",
        "怎么做更稳",
        "哪个更稳",
        "重启前是否",
        "要不要",
        "应不应该",
        "redundancy reload peer",
        "流量隔离",
        "隔离上下行",
        "隔离流量",
    ]

    return explicit_no_command or _batch67_contains_any(q, advice_keywords)


def _batch67_compact_context(context):
    if not isinstance(context, dict):
        return {}

    recent_turns = context.get("recent_turns")
    if not isinstance(recent_turns, list):
        recent_turns = []

    last_executions = context.get("last_executions")
    if not isinstance(last_executions, list):
        last_executions = []

    compact_exec = []
    for item in last_executions[-12:]:
        if not isinstance(item, dict):
            continue
        compact_exec.append({
            "command": item.get("command"),
            "status": item.get("status") or item.get("execution_status"),
            "ok": item.get("ok"),
            "analysis_status": item.get("analysis_status"),
            "analysis_summary": item.get("analysis_summary"),
            "output_preview": str(item.get("output_preview") or "")[:800],
        })

    return {
        "current_device": context.get("current_device"),
        "current_topic": context.get("current_topic"),
        "current_intent": context.get("current_intent"),
        "recent_turns_tail": recent_turns[-8:],
        "last_executions_tail": compact_exec,
        "last_inline_commands": context.get("last_inline_commands"),
    }


def _batch67_call_local_llm_for_advice(question, context):
    import json
    messages = [
        {
            "role": "system",
            "content": (
                "你是 NetAIOps 的网络运维方案建议助手。"
                "当前任务是回答用户的方案建议、操作建议、风险分析或优缺点比较问题。"
                "用户不是让你执行设备命令，也不是让你生成排查命令。"
                "你必须直接给出基于网络运维经验、当前会话上下文和已有结论的分析建议。"
                "不要编造已经执行过的命令结果；如果缺少现场证据，要明确说明不确定性。"
                "如果涉及高风险操作，例如重启、主备切换、流量隔离，要说明推荐顺序、风险点、回退条件和验证点。"
                "不要输出 Netmiko 待执行命令列表，不要要求用户输入 YES。"
            ),
        },
        {
            "role": "user",
            "content": (
                "当前会话上下文如下：\n"
                + json.dumps(_batch67_compact_context(context), ensure_ascii=False, indent=2)
                + "\n\n用户问题：\n"
                + str(question)
                + "\n\n请直接给出建议和分析。"
            ),
        },
    ]

    try:
        from netaiops_asset.chat_v2.llm_evidence_analyzer import _call_llm
        result = _call_llm(messages)
    except Exception as exc:
        return {
            "ok": False,
            "error": "local_llm_call_exception: %r" % (exc,),
            "content": "",
        }

    if not isinstance(result, dict):
        return {
            "ok": False,
            "error": "local_llm_return_not_dict: %r" % (type(result).__name__,),
            "content": "",
        }

    if result.get("ok") is False:
        return {
            "ok": False,
            "error": result.get("error") or result.get("message") or "local_llm_failed",
            "content": "",
            "raw": result,
        }

    content = (
        result.get("content")
        or result.get("answer")
        or result.get("text")
        or result.get("message")
        or ""
    )

    if not str(content).strip():
        return {
            "ok": False,
            "error": "local_llm_empty_content",
            "content": "",
            "raw": result,
        }

    return {
        "ok": True,
        "content": str(content).strip(),
        "raw": result,
    }


def _batch67_append_history(conversation_id, question, response, user):
    try:
        if "_batch66_sync_history_before_return" in globals():
            _batch66_sync_history_before_return(
                {
                    "question": question,
                    "user": user,
                    "conversation_id": conversation_id,
                },
                response,
            )
            return True
    except Exception:
        pass

    try:
        from netaiops_asset.agent.conversation_store import append_turn
        append_turn(conversation_id, question, response, user=user)
        return True
    except TypeError:
        try:
            from netaiops_asset.agent.conversation_store import append_turn
            append_turn(conversation_id, question, response)
            return True
        except Exception:
            return False
    except Exception:
        return False


def _batch67_save_v2_context(conversation_id, user, question, response):
    try:
        from netaiops_asset.chat_v2.context import save_v2_context_from_response
        save_v2_context_from_response(
            conversation_id=conversation_id,
            user=user,
            question=question,
            response=response,
        )
        return True
    except Exception as exc:
        try:
            response.setdefault("v2", {})["batch67_context_save_error"] = repr(exc)
        except Exception:
            pass
        return False


def _batch67_try_handle_advice_analysis(_locals):
    question = _batch67_extract_question_from_locals(_locals)
    if not _batch67_is_advice_analysis_question(question):
        return None

    user = _batch67_extract_user_from_locals(_locals)
    conversation_id = _batch67_extract_conversation_id_from_locals(_locals)

    try:
        from uuid import uuid4
        if not conversation_id:
            conversation_id = str(uuid4())
    except Exception:
        pass

    try:
        from netaiops_asset.chat_v2.context import load_v2_context
        context = load_v2_context(conversation_id=conversation_id, user=user) or {}
    except Exception:
        context = {}

    llm_result = _batch67_call_local_llm_for_advice(question, context)

    if llm_result.get("ok"):
        response = {
            "status": "ok",
            "planner_source": "v2_advice_analysis",
            "conversation_id": conversation_id,
            "answer": llm_result.get("content") or "",
            "items": [],
            "columns": [],
            "field_labels": {},
            "count": 0,
            "returned": 0,
            "parsed": {
                "intent": "v2_advice_analysis",
                "v2_intent": "advice_analysis",
                "reason": "batch67_advice_analysis_route",
                "requires_v2": True,
                "cmdb_only": False,
            },
            "llm_plan": {
                "action": "advice_analysis",
                "category": "operation_advice",
                "ok": True,
                "source": "batch67_advice_analysis_route",
            },
            "v2": {
                "batch67_advice_analysis": True,
                "no_command_generation": True,
                "no_device_execution": True,
                "context_device": context.get("current_device") if isinstance(context, dict) else None,
            },
        }
    else:
        response = {
            "status": "error",
            "planner_source": "v2_advice_analysis",
            "conversation_id": conversation_id,
            "answer": "本地 LLM 方案建议分析失败：%s" % (llm_result.get("error") or "unknown_error"),
            "items": [],
            "columns": [],
            "field_labels": {},
            "count": 0,
            "returned": 0,
            "parsed": {
                "intent": "v2_advice_analysis",
                "v2_intent": "advice_analysis",
                "reason": "batch67_advice_analysis_route_llm_failed",
            },
            "v2": {
                "batch67_advice_analysis": True,
                "no_command_generation": True,
                "no_device_execution": True,
                "llm_error": llm_result.get("error"),
            },
        }

    _batch67_save_v2_context(conversation_id, user, question, response)
    _batch67_append_history(conversation_id, question, response, user)

    return response


# ===== Batch68 conversation rename patch =====
import os as _batch68_os
import json as _batch68_json
import re as _batch68_re
from pathlib import Path as _Batch68Path
from fastapi import Request as _Batch68Request
from fastapi.responses import JSONResponse as _Batch68JSONResponse
from starlette.responses import Response as _Batch68Response

_BATCH68_TITLE_FILE = _Batch68Path(
    _batch68_os.getenv(
        "NETAIOPS_CONVERSATION_TITLE_FILE",
        "/var/lib/netaiops-asset-agent/data/conversation_titles.json",
    )
)

_BATCH68_SCRIPT_TAG = '<script src="/static/batch68_conversation_rename.js?v=batch68-rename" defer></script>'


def _batch68_title_store_read():
    try:
        if not _BATCH68_TITLE_FILE.exists():
            return {"version": 1, "users": {}}
        data = _batch68_json.loads(_BATCH68_TITLE_FILE.read_text(encoding="utf-8", errors="replace"))
        if not isinstance(data, dict):
            return {"version": 1, "users": {}}
        data.setdefault("version", 1)
        data.setdefault("users", {})
        if not isinstance(data.get("users"), dict):
            data["users"] = {}
        return data
    except Exception:
        return {"version": 1, "users": {}}


def _batch68_title_store_write(data):
    _BATCH68_TITLE_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = _BATCH68_TITLE_FILE.with_suffix(_BATCH68_TITLE_FILE.suffix + ".tmp")
    tmp.write_text(_batch68_json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(_BATCH68_TITLE_FILE)


def _batch68_sanitize_user(user):
    user = str(user or "web_user").strip()
    if not user:
        user = "web_user"
    return _batch68_re.sub(r"[^A-Za-z0-9_.@-]", "_", user)[:128]


def _batch68_sanitize_conversation_id(conversation_id):
    cid = str(conversation_id or "").strip()
    if not cid:
        return ""
    if len(cid) > 160:
        return ""
    if not _batch68_re.match(r"^[A-Za-z0-9_.:@-]+$", cid):
        return ""
    return cid


def _batch68_sanitize_title(title):
    title = str(title or "").replace("\r", " ").replace("\n", " ")
    title = _batch68_re.sub(r"\s+", " ", title).strip()
    if len(title) > 80:
        title = title[:80].strip()
    return title


def _batch68_get_user_from_request(request, payload=None):
    try:
        if isinstance(payload, dict) and payload.get("user"):
            return _batch68_sanitize_user(payload.get("user"))
    except Exception:
        pass
    try:
        q_user = request.query_params.get("user")
        if q_user:
            return _batch68_sanitize_user(q_user)
    except Exception:
        pass
    return "web_user"


@app.get("/api/v1/conversations/title-map")
async def batch68_get_conversation_title_map(request: _Batch68Request):
    user = _batch68_get_user_from_request(request)
    data = _batch68_title_store_read()
    titles = data.get("users", {}).get(user, {})
    if not isinstance(titles, dict):
        titles = {}
    return _Batch68JSONResponse({
        "ok": True,
        "user": user,
        "titles": titles,
        "count": len(titles),
    })


@app.patch("/api/v1/conversations/{conversation_id}/title")
@app.put("/api/v1/conversations/{conversation_id}/title")
@app.post("/api/v1/conversations/{conversation_id}/title")
async def batch68_set_conversation_title(conversation_id: str, request: _Batch68Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    if not isinstance(payload, dict):
        payload = {}

    user = _batch68_get_user_from_request(request, payload)
    cid = _batch68_sanitize_conversation_id(conversation_id)
    title = _batch68_sanitize_title(payload.get("title") or payload.get("name") or payload.get("custom_title"))

    if not cid:
        return _Batch68JSONResponse({
            "ok": False,
            "error": "invalid_conversation_id",
        }, status_code=400)

    if not title:
        return _Batch68JSONResponse({
            "ok": False,
            "error": "empty_title",
        }, status_code=400)

    data = _batch68_title_store_read()
    users = data.setdefault("users", {})
    user_titles = users.setdefault(user, {})
    if not isinstance(user_titles, dict):
        user_titles = {}
        users[user] = user_titles

    user_titles[cid] = title
    _batch68_title_store_write(data)

    return _Batch68JSONResponse({
        "ok": True,
        "user": user,
        "conversation_id": cid,
        "title": title,
        "title_file": str(_BATCH68_TITLE_FILE),
    })


@app.delete("/api/v1/conversations/{conversation_id}/title")
async def batch68_delete_conversation_title(conversation_id: str, request: _Batch68Request):
    user = _batch68_get_user_from_request(request)
    cid = _batch68_sanitize_conversation_id(conversation_id)

    if not cid:
        return _Batch68JSONResponse({
            "ok": False,
            "error": "invalid_conversation_id",
        }, status_code=400)

    data = _batch68_title_store_read()
    user_titles = data.get("users", {}).get(user, {})
    existed = False
    if isinstance(user_titles, dict) and cid in user_titles:
        existed = True
        user_titles.pop(cid, None)
        _batch68_title_store_write(data)

    return _Batch68JSONResponse({
        "ok": True,
        "user": user,
        "conversation_id": cid,
        "deleted": existed,
    })


@app.middleware("http")
async def batch68_conversation_rename_script_injector(request, call_next):
    response = await call_next(request)

    try:
        content_type = str(response.headers.get("content-type", ""))
        if "text/html" not in content_type.lower():
            return response

        body = b""
        async for chunk in response.body_iterator:
            body += chunk

        charset = "utf-8"
        html = body.decode(charset, errors="replace")

        if "batch68_conversation_rename.js" not in html:
            if "</body>" in html:
                html = html.replace("</body>", _BATCH68_SCRIPT_TAG + "\n</body>")
            else:
                html = html + "\n" + _BATCH68_SCRIPT_TAG + "\n"

        headers = dict(response.headers)
        headers.pop("content-length", None)
        return _Batch68Response(
            content=html,
            status_code=response.status_code,
            headers=headers,
            media_type="text/html",
        )
    except Exception:
        return response


# ===== Batch68-fix conversation title non-conflicting routes =====
@app.get("/api/v1/conversation-titles")
async def batch68_fix_get_conversation_titles(request: _Batch68Request):
    user = _batch68_get_user_from_request(request)
    data = _batch68_title_store_read()
    titles = data.get("users", {}).get(user, {})
    if not isinstance(titles, dict):
        titles = {}
    return _Batch68JSONResponse({
        "ok": True,
        "user": user,
        "titles": titles,
        "count": len(titles),
        "route": "/api/v1/conversation-titles",
        "batch68_fix": True,
    })


@app.patch("/api/v1/conversation-titles/{conversation_id}")
@app.put("/api/v1/conversation-titles/{conversation_id}")
@app.post("/api/v1/conversation-titles/{conversation_id}")
async def batch68_fix_set_conversation_title(conversation_id: str, request: _Batch68Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    if not isinstance(payload, dict):
        payload = {}

    user = _batch68_get_user_from_request(request, payload)
    cid = _batch68_sanitize_conversation_id(conversation_id)
    title = _batch68_sanitize_title(payload.get("title") or payload.get("name") or payload.get("custom_title"))

    if not cid:
        return _Batch68JSONResponse({
            "ok": False,
            "error": "invalid_conversation_id",
            "route": "/api/v1/conversation-titles/{conversation_id}",
            "batch68_fix": True,
        }, status_code=400)

    if not title:
        return _Batch68JSONResponse({
            "ok": False,
            "error": "empty_title",
            "route": "/api/v1/conversation-titles/{conversation_id}",
            "batch68_fix": True,
        }, status_code=400)

    data = _batch68_title_store_read()
    users = data.setdefault("users", {})
    user_titles = users.setdefault(user, {})
    if not isinstance(user_titles, dict):
        user_titles = {}
        users[user] = user_titles

    user_titles[cid] = title
    _batch68_title_store_write(data)

    return _Batch68JSONResponse({
        "ok": True,
        "user": user,
        "conversation_id": cid,
        "title": title,
        "title_file": str(_BATCH68_TITLE_FILE),
        "route": "/api/v1/conversation-titles/{conversation_id}",
        "batch68_fix": True,
    })


@app.delete("/api/v1/conversation-titles/{conversation_id}")
async def batch68_fix_delete_conversation_title(conversation_id: str, request: _Batch68Request):
    user = _batch68_get_user_from_request(request)
    cid = _batch68_sanitize_conversation_id(conversation_id)

    if not cid:
        return _Batch68JSONResponse({
            "ok": False,
            "error": "invalid_conversation_id",
            "route": "/api/v1/conversation-titles/{conversation_id}",
            "batch68_fix": True,
        }, status_code=400)

    data = _batch68_title_store_read()
    user_titles = data.get("users", {}).get(user, {})
    existed = False

    if isinstance(user_titles, dict) and cid in user_titles:
        existed = True
        user_titles.pop(cid, None)
        _batch68_title_store_write(data)

    return _Batch68JSONResponse({
        "ok": True,
        "user": user,
        "conversation_id": cid,
        "deleted": existed,
        "route": "/api/v1/conversation-titles/{conversation_id}",
        "batch68_fix": True,
    })
